
import os
import openpyxl
import csv
import reformat
import subprocess

from openpyxl import load_workbook


def gorilla_wtm_oocytes(ws):
    """
    Most up-to-date version of database importing. Reads oocyte data files in the format 'gorilla'

    This reads the oocyte data
    """
    wslist = []
    nonelist = [None] * 27
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=100, max_col=26):
        rowlist = []
        counter = 0
        for cell in row:
            if counter == 0:
                if cell.value == None:
                    break
                rowlist.append(reformat.create_expid(cell.value))
            rowlist.append(cell.value)
            counter = counter + 1
        if rowlist == nonelist:
            break
        else:
            wslist.append(rowlist)            
    header = ['expID','file','glun1','glun2','maxcurrent','ph68vs76','logm10','logm95','logm9','logm85',
              'logm8','logm75','logm7','logm65','logm6','logm55','logm5','logm45','logm4','logm35','logm3',
              'logm25','logm2','logm15','logm1','notes','dbinfo']
    wslist = [header] + wslist
    return wslist
def gorilla_wtm_experiments(ws):
    """
    Most up-to-date version of database importing. Reads oocyte data files in the format 'gorilla'

    This reads the experiment data
    """
    nonelist = [None] * 10
    for row in ws.iter_rows(min_row=3, min_col=27, max_row=3, max_col=36):
        file1 = ws['A3'].value
        experimentid = reformat.create_expid(file1)
        rowlist = [experimentid]
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist:
            break
    #header = ['experimentid','assay','date_inj','date_rec','vhold','coagonist','phsol','drugid','rig','initials','experiment_info']
    #wslist = [header] + wslist
    print(rowlist[0] + ' --- uploaded to experiments table in database')
    return rowlist
def turtle_wtm_experiments(ws):
    """
    Used for working with older cferv oocyte data recording sheets
    """
    wslist = []
    nonelist = [None] * 29
    rowcount = 0
    for row in ws.iter_rows(min_row=1, min_col=2, max_row=100, max_col=30):
        rowcount += 1
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist and wslist[rowcount-2] == nonelist:
            del wslist[rowcount-2]
            break
        else:
            wslist.append(rowlist)
    wslist = reformat.reformat(wslist, True)
    return wslist
def turtle_wtm_oocytes(ws):
    """
    Used for working with older cferv oocyte data recording sheets
    """
    wslist = []
    nonelist = [None] * 29
    rowcount = 0
    for row in ws.iter_rows(min_row=1, min_col=2, max_row=100, max_col=30):
        rowcount += 1
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist and wslist[rowcount-2] == nonelist:
            del wslist[rowcount-2]
            break
        else:
            wslist.append(rowlist)
    wslist = reformat.reformat(wslist, False)
    return wslist
def fox_wtm(ws):
    """
    Reads oocyte data files in the format 'fox'. 
    
    This refers to files created in the early days of database developement (7/17/18 - 8/17/18)
    The files have been significantly changed since then.
    """
    wslist = []
    nonelist = [None] * 37
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=100, max_col=37):
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist:
            break
        else:
            wslist.append(rowlist)
    header = ['assay','file','glun1','glun2','maxcurrent','ph68vs76','logm10','logm95','logm9',
              'logm85','logm8','logm75','logm7','logm65','logm6','logm55','logm5','logm45','logm4',
              'logm35','logm3','logm25','logm2','logm15','logm1','notes','date_inj','date_rec',
              'vhold','coagonist','phsol','rig','initials','dpi','rnapotglun1','rnapotglun2','dbinfo']
    wslist = [header] + wslist
    return wslist
def read_csv(csvfile):
    with open(csvfile, mode = 'r') as fp:
        reader = csv.reader(fp, delimiter=',', quotechar='"')
        data_read = [row for row in reader]
    return(data_read)
def write_csv(filename, csvname, wtm_function): # you should specify which wtm function you use (worksheet to matrix)
    wb = openpyxl.load_workbook(filename, data_only = True)
    ws = wb.worksheets[0]
    wslist = wtm_function(ws)
    #wtm_function chosen based on arguments
    with open(csvname, 'w', newline = '') as csvfile:
        testwriter = csv.writer(csvfile)
        for wslist in wslist:
            testwriter.writerow(wslist)
        return(None)
    return(None)
def excel_to_csv(readdir, writedir, wtm_function): #writes all excel files in a folder to a csv in a folder
    for filename in os.listdir(readdir):
        csvfilename = filename[0:(len(filename)-4)] + 'csv'
        readfulldir = readdir + '/' + filename
        writefulldir = writedir + '/' + csvfilename
        write_csv(readfulldir, writefulldir, wtm_function)
        #wtm_function chosen based on arguments
        print(csvfilename + ' --- written to desktop folder csvdump')
    writelen = len(os.listdir(readdir))
    writestr = str(writelen) + " files written to csv"
    print(writestr)
    return(None)
def excel_to_db(readdir, wtm_function, db_function):
    for filename in os.listdir(readdir):
        readfulldir = readdir + '/' + filename
        wb = load_workbook(readfulldir, data_only = True)
        ws = wb.worksheets[0]
        db_function(wtm_function(ws))
    writelen = len(os.listdir(readdir))
    writestr = str(writelen) + " files uploaded to database"
    print(writestr)
    return None
def csv_to_db(readdir, db_function):
    count = 0
    for filename in os.listdir(readdir):
        readfulldir = readdir + '/' + filename
        print(readfulldir)
        filelist = read_csv(readfulldir)
        code = db_function(filelist)
        count += code
    print(str(count) + " file(s) uploaded to database")
    return None
def call_rscript(readdir, scriptpath): #exectues an Rscript as a subprocess
    command = 'Rscript'
    for filename in os.listdir(readdir):
        filefull = readdir + '/' + filename
        if filename[11:13] == 'Zn':
            bot = 'FALSE'
            top = 'TRUE'
        else:
            bot = 'TRUE'
            top = 'TRUE' 
        cmd = [command , scriptpath , filefull, bot, top]
        subprocess.call(cmd)
        print(filefull + '--- added fits to the file')
    return(None)
def iterexcel(readdir, wtm_function):
    for filename in os.listdir(readdir):
        readfulldir = readdir + '/' + filename
        wb = load_workbook(readfulldir, data_only = True)
        ws = wb.worksheets[0]
        wslist = wtm_function(ws)
        for row in wslist:
            print(row)
        #listfunction(wslist)
    return wslist
