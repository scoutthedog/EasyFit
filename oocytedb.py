#! C:/Users/jpa/Anaconda3/python

"""
For processing and uploading oocyte recording data stored in excel files into a MySQL database

A general outline of the flow of this program.
1) Uses openpyxl to read excel files and import them into python lists.
2) The lists are then written to csv files
3) The csv files are read by an R subprocess which calls Rscript and passes arguments from commandline
4) The Rscript generates fits for the data and rewrites the csv file including the information about the fits
5) The now fit data is uploaded to the MySQL database
"""

import os
import csv
import subprocess
import mysql.connector
import pandas
import datetime
from openpyxl import load_workbook

RECORDING_FILES_DIRECTORY = 'C:/Users/jpa/Desktop/uploaddata'
CSV_WRITE_DIRECTORY = 'C:/Users/jpa/Desktop/csvdump'
RSCRIPT_PATH = 'C:/Users/jpa/source/repos/oocytedb/oocytescript.R'
DATABASE = 'oocytedb'

class excelcoordinates:
    """
    This class is used in order to convert older spreadsheet files for database entry.

    On class initiation, the coordinates are entered as tuples. A matrix of values derrived straight from
    the excel spreadsheet are inputted and then it returns a correctly formatted matrix.
    """
    def __init__(self, assay, postp, date_inj, date_rec, vhold,
                coagonist, phsol, drugid, rig, 
                initials, note, filename, glun1, 
                glun2, current, d_start,
               d_end, r_start, r_end, rec_note):
        """
        The Class is initialized with the coordinates as tuples. assay and postp are the only non tuple arguments
        """
        self.assay = assay #assay string
        self.postp = postp #post processing function used for the class instance
        
        #self.experiments : contains a dictionary of tuples as coordinates that refer to their location in the wtm_function
        #wtm_function -> raw data as a list-> excelcoordinates -> postp -> ready for database upload
        self.experiments = ({'date_inj':date_inj, 'date_rec':date_rec, 'vhold':vhold, 
                             'coagonist':coagonist, 'phsol':phsol, 'drugid':drugid, 'rig':rig, 'initials':initials, 
                             'note':note})
        #Some aberviations:
        #d_start - dose start
        #d_end - dose end - these are the headers for the data, determines where the data should be placed
        #r_start - response start
        #r_end - response end
        #self.oocytes : also a dictionary of tuples that refer to the locations that the data is found on the excel spreadsheet.
        self.oocytes = ({'filename':filename, 'glun1':glun1, 
                             'glun2':glun2, 'current':current, 'd_start':d_start, 
                             'd_end':d_end, 'r_start':r_start, 'r_end':r_end, 'rec_note':rec_note})
    def getexperiments(self, wslist):
        #leaves the first value expID blank, and sets assay to what was given at initialization
        experimentlist = [None, self.assay]
        #iterating through keys and value of the experiments dictionary
        for key, value in self.experiments.items():
            if value == None:
                #checking if the dictionary value is None
                #value stays None
                wsvalue = value
            else:
                #value is set to whatever is in the coordinates of wslist
                x = value[0]
                y = value[1]
                wsvalue = wslist[x][y]
            experimentlist.append(wsvalue)
        #postp processes the values in the experiment list (see postp functions)
        return(self.postp(experimentlist))
    def getoocytes(self, wslist):
        #d is now the dictionary of oocytes coordinates
        d = self.oocytes
        #an integer specifing the first row containing data (usually row 6 or 7)
        rowstart = d['filename'][0]
        #an integer specified the row containing the dose/concentration
        dosestart = d['d_start'][0]
        #using its own method getexperiments()
        explist = self.getexperiments(wslist)
        #experiment id will be used to calculate the filnames
        expid = explist[0]
        #left blank
        oocytes = []
        #behavior is different for pH
        if self.assay == 'pH':
            for row in range(rowstart, len(wslist)):
                filename = wslist[row][d['filename'][1]]
                k = filename.rfind('-')
                filename = expid + filename[k:]
                note = wslist[row][d['rec_note'][1]]
                nonelist = [None] * 17
                rowlist = [filename] + [wslist[row][d['glun1'][1]], wslist[row][d['glun2'][1]], 
                                        wslist[row][d['current'][1]], wslist[row][d['r_start'][1]]] + nonelist + [wslist[row][d['rec_note'][1]]] + [None]
                oocytes.append(rowlist)
        #all other assays
        else:
            #the header is just a temporary placeholder that will be used to determine where to place the data
            header = ['filename','glun1','glun2','current']
            for dose in range(d['d_start'][1], d['d_end'][1] + 1):
                heading = wslist[dosestart][dose]
                header.append(heading)
            doses = header[4:]
            #this code is designed to figure out how many 'none' values need to be added before and after...
            #... in order to properly align the data for database entry.
            fulldose = [-10, -9.52, -9, -8.52, -8, -7.52, -7, -6.52, -6, -5.52, -5, -4.52, -4, -3.52, -3, -2.52, -2, -1.52, -1]
            rounded = []
            for value in doses:
                if value == None:
                    pass
                else:
                    value = round(value, 2)
                rounded.append(value)
            indexes = []
            for value in rounded:
                value = fulldose.index(value)
                indexes.append(value)
            #cal
            nonecolspre = [None] * min(indexes)
            nonecolspost = [None] * (18 - max(indexes))
            for row in range(rowstart, len(wslist)):
                #creating the filename
                filename = wslist[row][d['filename'][1]]
                k = filename.rfind('-')
                filename = expid + filename[k:]
                note = wslist[row][d['rec_note'][1]]
                rowlist = [filename] + [wslist[row][d['glun1'][1]], wslist[row][d['glun2'][1]], wslist[row][d['current'][1]]] + [None] + nonecolspre
                for response in range(d['r_start'][1], d['r_end'][1] + 1):
                    record = wslist[row][response]
                    rowlist.append(record)
                rowlist = rowlist + nonecolspost + [note] + [None]
                oocytes.append(rowlist)
        return(oocytes)

#--------------------
# DATABASE FUNCTIONS |
#--------------------
def dbcon():
    """
    opens a new MySQL database connection
    """
    con = mysql.connector.connect(user='root', host='localhost', database=DATABASE)
    return con
def experiments_update(explist):
    """
    Connects to a MySQL database and uploads a list of experiment data to the table 'experiments'
    """
    if explist == []:
        return(None)
    con = dbcon()
    cursor = con.cursor()
    query = ("REPLACE INTO experiments"
             "(expID, assay, date_inj, date_rec, vhold, coagonist, phsol, drugID, rig, initials, experiment_info)"
             "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
             )
    cursor.execute(query, explist)
    con.commit()
    cursor.close()
    con.close
    return(None)
def oocytes_update(oocytelist):
    """
    Connects to a MySQL database and uploads a maxtrix (2-dimentional list) of data to the table 'oocytes'

    Prints data on command line, and the user confirms that the data is accurate before it is uploaded.
    The MySQL query uses the REPLACE syntax which means that old data is overwritten if the primary key (file)
    is already found in the 'oocytes' table.
    """
    con = dbcon()
    cursor = con.cursor()
    #pandas dataframe used to format the matrix nicely
    print(pandas.DataFrame(oocytelist).to_string())
    for sublist in oocytelist:
        query = ("REPLACE INTO oocytes" 
                 "(expID,file,glun1,glun2,maxcurrent,ph68vs76,"
                 "logm10,logm95,logm9,logm85,logm8,logm75,logm7,logm65,logm6,logm55,logm5,logm45,logm4,logm35,"
                 "logm3,logm25,logm2,logm15,logm1,notes,dbinfo,logec50,hillslope,ymin,ymax,formula,isConv)"
                 "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, "
                 "%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                 )
        cursor.execute(query, sublist)
    isgood = input("Does this look good? (y/n) ")
    if isgood == "y":
        con.commit()
        print("Uploaded to database")
        code = 1
    else:
        con.close
        print("Did not upload")
        code = 0
    con.close
    # code = 1 : uploaded
    # code = 0 : not uploaded
    return code
#---------------------
# WORKSHEET FUNCTIONS|
#---------------------
def fox_wtm(ws):
    """
    Reads oocyte data files in the format 'fox'. 
    
    This refers to files created in the early days of database developement (7/17/18 - 8/17/18)
    The files have been significantly changed since then.
    """
    wslist = []
    nonelist = [None] * 37
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=100, max_col=37):
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist:
            break
        else:
            wslist.append(rowlist)
    header = ['assay','file','glun1','glun2','maxcurrent','ph68vs76','logm10','logm95','logm9',
              'logm85','logm8','logm75','logm7','logm65','logm6','logm55','logm5','logm45','logm4',
              'logm35','logm3','logm25','logm2','logm15','logm1','notes','date_inj','date_rec',
              'vhold','coagonist','phsol','rig','initials','dpi','rnapotglun1','rnapotglun2','dbinfo']
    wslist = [header] + wslist
    return wslist
def gorilla_wtm_oocytes(ws):
    """
    Most up-to-date version of database importing. Reads oocyte data files in the format 'gorilla'

    This reads the oocyte data
    """
    wslist = []
    nonelist = [None] * 27
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=100, max_col=26):
        rowlist = []
        counter = 0
        for cell in row:
            if counter == 0:
                if cell.value == None:
                    break
                rowlist.append(create_expid(cell.value))
            rowlist.append(cell.value)
            counter = counter + 1
        if rowlist == nonelist:
            break
        else:
            wslist.append(rowlist)            
    header = ['expID','file','glun1','glun2','maxcurrent','ph68vs76','logm10','logm95','logm9','logm85',
              'logm8','logm75','logm7','logm65','logm6','logm55','logm5','logm45','logm4','logm35','logm3',
              'logm25','logm2','logm15','logm1','notes','dbinfo']
    wslist = [header] + wslist
    return wslist
def gorilla_wtm_experiments(ws):
    """
    Most up-to-date version of database importing. Reads oocyte data files in the format 'gorilla'

    This reads the experiment data
    """
    nonelist = [None] * 10
    for row in ws.iter_rows(min_row=3, min_col=27, max_row=3, max_col=36):
        file1 = ws['A3'].value
        experimentid = create_expid(file1)
        rowlist = [experimentid]
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist:
            break
    #header = ['experimentid','assay','date_inj','date_rec','vhold','coagonist','phsol','drugid','rig','initials','experiment_info']
    #wslist = [header] + wslist
    print(rowlist[0] + ' --- uploaded to experiments table in database')
    return rowlist
def turtle_wtm_experiments(ws):
    """
    Used for working with older cferv oocyte data recording sheets
    """
    wslist = []
    nonelist = [None] * 29
    rowcount = 0
    for row in ws.iter_rows(min_row=1, min_col=2, max_row=100, max_col=30):
        rowcount += 1
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist and wslist[rowcount-2] == nonelist:
            del wslist[rowcount-2]
            break
        else:
            wslist.append(rowlist)
    wslist = reformat(wslist, True)
    print(wslist)
    return wslist
def turtle_wtm_oocytes(ws):
    """
    Used for working with older cferv oocyte data recording sheets
    """
    wslist = []
    nonelist = [None] * 29
    rowcount = 0
    for row in ws.iter_rows(min_row=1, min_col=2, max_row=100, max_col=30):
        rowcount += 1
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        if rowlist == nonelist and wslist[rowcount-2] == nonelist:
            del wslist[rowcount-2]
            break
        else:
            wslist.append(rowlist)
    wslist = reformat(wslist, False)
    return wslist

def postpA(explist):
    """
    This is a post processing function A. It is used by the reformat function

    Mutiple post-processing functions may be added with developement.
    """
    date_inj = explist[2]
    if type(date_inj) != datetime.datetime:
        date_inj = None
    explist[2] = date_inj
    vhold = explist[4]
    vhold = vhold.strip('Vhold=')
    vhold = vhold.strip(' mV')
    explist[4] = vhold
    phsol = explist[6]
    phsol = phsol.strip('pH ')
    explist[6] = phsol
    rig = explist[8]
    rig = explist[8].replace('#','')
    initials = explist[9]
    if initials == 'skim':
        initials = 'sk'
    explist[9] = initials
    minidate = explist[3].strftime('%m%d%y')
    miniassay = explist[1].strip('DRC')
    explist[0] = miniassay + '-' + str(rig) + '-' + initials + '-' + minidate
    return(explist)
def reformat(wslist, exp):
    """
    wslist generated by the wtm function (raw unedited data from excel) --> database ready list

    The bulk of this function workhorse is the excelcoordinates class. (defined below)
    When an excelcoordinates class is defined, it is given tuples to describe the listwise location of the data.
    For an example:
    date_rec = (1,3) --> wslist[1][3] 
    Whatever is stored in wslist[1][3] when given to this function,
    will eventually be imported into the database as the date_rec.
    This is why it is important to know where the data is stored on each excel spread sheet.
    This function will (attemp to) logically determine which class instance should be used
    As different formats are encountered in old data, this function will be modified
    Additional class instances will be added as well
    """
    assay = wslist[2][5]
    if assay == 'Glycine DRC (uM)':
        assay = 'glyDRC'
    elif assay == 'Glutamate DRC (uM)':
        assay = 'gluDRC'
    elif assay == 'pH':
        assay = 'pH'
    elif assay == '1 min 1 conc ':
        assay = '1min1conc'
    elif assay == None:
        assay = wslist[2][4]
        if assay == 'Mg2+ (uM)':
            assay = 'mgDRC'
    else:
        assay = wslist[0][3]
        if assay == 'Zn2+ inhibition (10 mM Tricine)':
            assay = 'znDRC'
        else:
            assay = wslist[1][3]
            if assay == 'in 100 uM Glutamate and 100uM Glycine':
                assay = 'mGluGly'
    # CLASS INSTANCES
    glu = excelcoordinates(assay = 'gluDRC', postp = postpA, date_inj = (0,1), date_rec = (1,1), vhold = (0,3), coagonist = (1,3), phsol = (0,5), drugid = None,
                       initials = (0,7), rig = (1,7), note = (0,9), filename = (6,0), glun1 = (7,1), glun2 = (7,2), 
                       current = (5,3), d_start = (4,5), d_end = (4,14), r_start = (7,5), r_end = (7,14), rec_note = (7,15))
    gly = excelcoordinates(assay = 'glyDRC', postp = postpA, date_inj = (0,1), date_rec = (1,1), vhold = (0,3), coagonist = (1,3), phsol = (0,5), drugid = None,
                       initials = (0,7), rig = (1,7), note = (0,9), filename = (6,0), glun1 = (7,1), glun2 = (7,2), 
                       current = (5,3), d_start = (4,5), d_end = (4,14), r_start = (7,5), r_end = (7,14), rec_note = (7,15))
    mg = excelcoordinates(assay = 'mgDRC', postp = postpA, date_inj = (0,1), date_rec = (1,1), vhold = (0,5), coagonist = (1,3), phsol = (1,5), drugid = None,
                      initials = (0,8), rig = (1,8), note = (0,10), filename = (6,0), glun1 = (7,1), glun2 = (7,2),
                      current = (5,3), d_start = (4,5), d_end = (4,13), r_start = (7,5), r_end = (7,13), rec_note = (7,14))
    ph = excelcoordinates(assay = 'pH', postp = postpA, date_inj = (0,1), date_rec = (1,1), vhold = (0,5), coagonist = (1,2), phsol = (0,2), drugid = None,
                      initials = (0,9), rig = (1,9), note = (0,11), filename = (6,0), glun1 = (7,1), glun2 = (7,2),
                      current = (7,3), d_start = (4,6), d_end = (4,6), r_start = (7,6), r_end = (7,6), rec_note = (7,15))
    zn = excelcoordinates(assay = 'znDRC', postp = postpA, date_inj = (0,1), date_rec = (1,1), vhold = (0,5), coagonist = (1,3), phsol = (1,5), drugid = None,
                      initials = (0,7), rig = (1,7), note = (0,9), filename = (6,0), glun1 = (5,1), glun2 = (5,2),
                      current = (5,3), d_start = (4,5), d_end = (4,10), r_start = (7,5), r_end = (7,10), rec_note = (7,14))
    # Class selection logic
    if assay == 'gluDRC':
        explist = glu.getexperiments(wslist)
        oocytes = glu.getoocytes(wslist)
    elif assay == 'glyDRC':
        explist = gly.getexperiments(wslist)
        oocytes = gly.getoocytes(wslist)
    elif assay == 'mgDRC':
        explist = mg.getexperiments(wslist)
        oocytes = mg.getoocytes(wslist)
    elif assay == 'znDRC':
        explist = zn.getexperiments(wslist)
        oocytes = zn.getoocytes(wslist)
    elif assay == 'pH':
        explist = ph.getexperiments(wslist)
        oocytes = ph.getoocytes(wslist)
    else:
        explist = []
        oocytes = []
    if exp == True:
        return explist
    else:
        return oocytes

#-------------------------------------------------------------------------------------------------------------
# CLASSES


#--------------------------------------------------------------------------------------------------

#------------------------------------------------------------------------------------------------
# OTHER FUNCTIONS

def find_nth(mystring, x, n, i = 0):
    """
    find the nth occurance of string 'x' in mystring
    """
    i = mystring.find(x, i)
    if n == 1 or i == -1:
        return i 
    else:
        return find_nth(mystring, x, n - 1, i + len(x))
def create_expid(file):
    """
    Function for creating the experiment ID. 
    
    This is an important primary key in the experiments table of the database,
    and a foriegn key to the oocytes table.
    """
    test = find_nth(file, '-', 4)
    #find the 4th occurance of '-' in file
    expid = file[:test]
    #remove anything after the 4th occurance of '-'
    #example: glu-mb1-jpa-082918-14.5   -->  glu-mb1-jpa-082918
    return expid
#-------------------------------------------------------------------------------------------------
# DIRECTORY FUNCTIONS
# these functions works with files on your hard drive
# main data formats: csv, excel, MySQL
def read_csv(csvfile):
    with open(csvfile, mode = 'r') as fp:
        reader = csv.reader(fp, delimiter=',', quotechar='"')
        data_read = [row for row in reader]
    return(data_read)
def write_csv(filename, csvname, wtm_function): # you should specify which wtm function you use (worksheet to matrix)
    wb = load_workbook(filename, data_only = True)
    ws = wb.worksheets[0]
    wslist = wtm_function(ws)
    #wtm_function chosen based on arguments
    with open(csvname, 'w', newline = '') as csvfile:
        testwriter = csv.writer(csvfile)
        for wslist in wslist:
            testwriter.writerow(wslist)
        return(None)
    return(None)
def excel_to_csv(readdir, writedir, wtm_function): #writes all excel files in a folder to a csv in a folder
    for filename in os.listdir(readdir):
        csvfilename = filename[0:(len(filename)-4)] + 'csv'
        readfulldir = readdir + '/' + filename
        writefulldir = writedir + '/' + csvfilename
        write_csv(readfulldir, writefulldir, wtm_function)
        #wtm_function chosen based on arguments
        print(csvfilename + ' --- written to desktop folder csvdump')
    writelen = len(os.listdir(readdir))
    writestr = str(writelen) + " files written to csv"
    print(writestr)
    return(None)
def excel_to_db(readdir, wtm_function, db_function):
    for filename in os.listdir(readdir):
        readfulldir = readdir + '/' + filename
        wb = load_workbook(readfulldir, data_only = True)
        ws = wb.worksheets[0]
        db_function(wtm_function(ws))
    writelen = len(os.listdir(readdir))
    writestr = str(writelen) + " files uploaded to database"
    print(writestr)
    return None
def csv_to_db(readdir, db_function):
    count = 0
    for filename in os.listdir(readdir):
        readfulldir = readdir + '/' + filename
        print(readfulldir)
        filelist = read_csv(readfulldir)
        code = db_function(filelist)
        count += code
    print(str(count) + " file(s) uploaded to database")
    return None
def call_rscript(readdir, scriptpath): #exectues an Rscript as a subprocess
    command = 'Rscript'
    for filename in os.listdir(readdir):
        filefull = readdir + '/' + filename
        if filename[11:13] == 'Zn':
            bot = 'FALSE'
            top = 'TRUE'
        else:
            bot = 'TRUE'
            top = 'TRUE' 
        cmd = [command , scriptpath , filefull, bot, top]
        subprocess.call(cmd)
        print(filefull + '--- added fits to the file')
    return(None)

# Primary function.
def dbupload(wtm_function_experiments, wtm_function_oocytes): # This is the main script function  
    
    print('\nInitiating csv creation for R subprocess\n')
    excel_to_csv(RECORDING_FILES_DIRECTORY, CSV_WRITE_DIRECTORY, wtm_function_oocytes) 
    print('\nInitiating R curve fitting\n')
    call_rscript(CSV_WRITE_DIRECTORY, RSCRIPT_PATH)
    print('\nCurves fit. Preparing for database upload\n')
    csv_to_db(CSV_WRITE_DIRECTORY, oocytes_update)
    print('Initiating experiments data upload\n')
    excel_to_db(RECORDING_FILES_DIRECTORY, wtm_function_experiments, experiments_update) 
    return None
#testing
def iterexcel(readdir, wtm_function, listfunction):
    for filename in os.listdir(readdir):
        readfulldir = readdir + '/' + filename
        wb = load_workbook(readfulldir, data_only = True)
        ws = wb.worksheets[0]
        wslist = wtm_function(ws)
        listfunction(wslist)
    return wslist

#---------------------------------------------------------------------------
#MAIN
def main():
    dbupload(gorilla_wtm_experiments, gorilla_wtm_oocytes)
    return None
if __name__ == "__main__":
    main()


